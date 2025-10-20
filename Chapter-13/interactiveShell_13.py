# Chapter 13 - WORKING WITH EXCEL SPREADSHEETS - Automate the boring stuff

## Reading Excel Document

## -- Opening Excel Documents with OpenPyXl --

import openpyxl

"""
wb = openpyxl.load_workbook("example.xlsx")
print(type(wb))
"""

## -- Getting Sheets from the Workbook --

"""
wb = openpyxl.load_workbook("example.xlsx")
print(wb.sheetnames)  # The workbook's sheet names.
sheet = wb["Sheet3"]  # Get a sheet from the workbook.
print(sheet)
print(type(sheet))
print(sheet.title)  # Get the sheet's title as a string.
anotherSheet = wb.active  # Get the active sheet.
print(anotherSheet)
"""

## -- Getting Cells from the Sheet --

"""
wb = openpyxl.load_workbook("example.xlsx")
sheet = wb["Sheet1"]  # Get a sheet from the workbook.
print(sheet["A1"])  # Get a cell from the sheet.
print(sheet["A1"].value)  # Get the value from the cell.
c = sheet["B1"]  # Get another cell from the sheet.
print(c.value)
# Get the row, column, and value from the cell.
print("Row %s, Column %s is %s" % (c.row, c.column, c.value))
print("Cell %s is %s" % (c.coordinate, c.value))
print(sheet["C1"].value)

print(sheet.cell(row=1, column=2))
print(sheet.cell(row=1, column=2).value)

for i in range(1, 8, 2):  # Go through every other row:
    print(i, sheet.cell(row=i, column=2).value)
"""

"""
wb = openpyxl.load_workbook("example.xlsx")
sheet = wb["Sheet1"]
print(sheet.max_row)  # Get the highest row number
print(sheet.max_column)  # Get the highest column number
"""

## -- Converting Between Column Letter and Numbers

from openpyxl.utils import get_column_letter, column_index_from_string

"""
print(get_column_letter(1))  # Translate column 1 to a letter.
print(get_column_letter(2))
print(get_column_letter(27))
print(get_column_letter(900))

wb = openpyxl.load_workbook("example.xlsx")
sheet = wb["Sheet1"]
print(get_column_letter(sheet.max_column))
print(column_index_from_string("A"))  # Get A's number.
print(column_index_from_string("AA"))
"""

## -- Getting Rows ans Columns from the Sheets --

"""
wb = openpyxl.load_workbook("example.xlsx")
sheet = wb["Sheet1"]
print(tuple(sheet["A1":"C3"]))  # Get all cells from A1 to C3.
for rowOfCellObjects in sheet["A1":"C3"]:
    for cellObj in rowOfCellObjects:
        print(cellObj.coordinate, cellObj.value)
    print("... END OF ROW ...")
"""

"""
wb = openpyxl.load_workbook("example.xlsx")
sheet = wb.active
print(list(sheet.columns)[1])  # Get the second column's cells.
for cellObj in list(sheet.columns)[1]:
    print(cellObj.value)
"""

## -- Workbooks, Sheets, Cells --

"""
import os
import census2010

print(census2010.allData["AK"]["Anchorage"])
anchoragePop = census2010.allData["AK"]["Anchorage"]["pop"]
print("The 2010 population of Anchorage was " + str(anchoragePop))
"""

## Writing Excel Documents

## -- Crating and Saving Excel Documents --
